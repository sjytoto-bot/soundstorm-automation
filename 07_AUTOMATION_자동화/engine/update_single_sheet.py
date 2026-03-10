#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
사용자 마스터 레이아웃 보존형 데이터 동기화 엔진 (Safe Sync)
- 기존 서식(배경색, 테두리, 병합, 조건부 서식)을 건드리지 않고 데이터만 업데이트
"""

import json
import openpyxl
from pathlib import Path
from google.oauth2 import service_account
from googleapiclient.discovery import build

BASE_DIR = Path(__file__).parent
CREDENTIALS_DIR = BASE_DIR / 'credentials'
SERVICE_ACCOUNT_FILE = CREDENTIALS_DIR / 'service_account.json'
SCOPES = ['https://www.googleapis.com/auth/spreadsheets']

def get_sheets_service():
    creds = service_account.Credentials.from_service_account_file(
        SERVICE_ACCOUNT_FILE, scopes=SCOPES)
    return build('sheets', 'v4', credentials=creds)

def update_master_data_only():
    """사용자 마스터 파일의 레이아웃을 유지하며 데이터만 업데이트"""
    
    # 1. 최신 엑셀 파일 찾기
    excel_files = sorted(BASE_DIR.glob('YouTube_종합인사이트_리포트_*.xlsx'), reverse=True)
    if not excel_files:
        print("❌ 엑셀 파일을 찾을 수 없습니다.")
        return
    
    excel_file = excel_files[0]
    print(f"📂 최신 리포트 로드: {excel_file.name}")
    
    # 2. 구글 시트 ID 로드
    config_file = BASE_DIR / 'google_sheets_config.json'
    with open(config_file, 'r', encoding='utf-8') as f:
        spreadsheet_id = json.load(f).get('spreadsheet_id')
    
    # 3. 데이터 동기화 서비스 생성
    service = get_sheets_service()
    
    # 시트 목록 조회
    spreadsheet = service.spreadsheets().get(spreadsheetId=spreadsheet_id, includeGridData=False).execute()
    spreadsheet_titles = [s['properties']['title'] for s in spreadsheet['sheets']]
    
    # 엑셀 워크북 로드
    wb = openpyxl.load_workbook(excel_file, data_only=True)

    # --- [함수: 블록 업데이트 및 정제] ---
    def update_block(sheet_name, start_cell, excel_sheet_name, excel_start_row, excel_start_col, rows, cols, clean_gender=False):
        """엑셀의 특정 블록을 구글 시트의 시작 셀부터 채워넣음 (데이터 정제 포함)"""
        if excel_sheet_name not in wb.sheetnames:
            return
        ws_excel = wb[excel_sheet_name]
        data = []
        for r in range(excel_start_row, excel_start_row + rows):
            row_data = []
            for c in range(excel_start_col, excel_start_col + cols):
                val = ws_excel.cell(row=r, column=c).value
                val = '' if val is None else val
                
                # 성별 데이터 정제 (여성, 남성, 기타)
                if clean_gender and c == (excel_start_col + 1):
                    gender_map = {'female': '여성', 'male': '남성', 'genderUserSpecified': '기타'}
                    val = gender_map.get(val, val)
                
                row_data.append(val)
            data.append(row_data)
        
        service.spreadsheets().values().update(
            spreadsheetId=spreadsheet_id,
            range=f"'{sheet_name}'!{start_cell}",
            valueInputOption='USER_ENTERED',
            body={'values': data}
        ).execute()
        print(f"  ✅ {sheet_name} > {start_cell} 블록 완료")

    # --- [섹션 1: (종합) 전체기간 분석] ---
    sheet_name = '(종합) 전체기간 분석'
    if sheet_name in spreadsheet_titles:
        print(f"🔄 '{sheet_name}' 레이아웃 복구 및 최종 정밀 매핑 중...")
        
        # 1. 인구 및 국가 헤더 복구 (Row 12)
        # Excel 11행의 헤더(연령, 성별, 비중...)를 GSheet 12행에 복구
        update_block(sheet_name, "B12", sheet_name, 11, 1, 1, 3) # 연령/성별 헤더
        update_block(sheet_name, "F12", sheet_name, 11, 6, 1, 3) # 국가 헤더
        
        # 2. 데이터 업데이트 (Row 13부터)
        update_block(sheet_name, "B13", sheet_name, 12, 1, 15, 3, clean_gender=True) # 인구 데이터
        update_block(sheet_name, "F13", sheet_name, 12, 6, 15, 3) # 국가 데이터
        
        # 3. 검색어 섹션 복구 (Row 26~28)
        update_block(sheet_name, "B26", sheet_name, 26, 1, 1, 1) # 타이틀
        update_block(sheet_name, "B27", sheet_name, 27, 1, 1, 3) # 헤더 복구
        update_block(sheet_name, "B28", sheet_name, 28, 1, 15, 3) # 검색어 데이터
        
        # 4. 기기별 비중 복구 (Row 44~46)
        # 유령 데이터 44행 청소 후 헤더 및 데이터 배치
        update_block(sheet_name, "F44", sheet_name, 45, 6, 1, 3) # 기기 헤더
        update_block(sheet_name, "F45", sheet_name, 46, 6, 10, 3) # 기기 데이터

        # 🧹 유령 데이터 및 잔여물 정밀 청소
        # 1. 44행의 과거 잔여물 2. F28:H43 부근의 중복 데이터 3. 하단 60행 이후 영역
        clear_ranges = [
            f"'{sheet_name}'!A44:E44", 
            f"'{sheet_name}'!F26:H43", # 중복된 기기 데이터 등
            f"'{sheet_name}'!A60:Z150"  # 하단 여백 정리
        ] 
        for cr in clear_ranges:
            service.spreadsheets().values().clear(spreadsheetId=spreadsheet_id, range=cr).execute()
        print(f"  🧹 '{sheet_name}' 최종 미세 청소 완료")

    # --- [섹션 2: 트랜드분석&인사이트] ---
    sheet_trend = '트랜드분석&인사이트'
    excel_guide = '(가이드) 트렌드 & 패턴'
    
    if sheet_trend in spreadsheet_titles and excel_guide in wb.sheetnames:
        print(f"🔄 '{sheet_trend}' 정밀 매핑 중...")
        # 엑셀 (가이드) 시트의 성과 트렌드 비교 표 (행 4~7) -> 구글 시트 B7
        update_block(sheet_trend, "B7", excel_guide, 4, 1, 4, 4)

    print(f"\n🚀 사용자 마스터 파일 데이터 업데이트 완료 (Safe Sync)!")
    print(f"🔗 https://docs.google.com/spreadsheets/d/{spreadsheet_id}/edit")

if __name__ == "__main__":
    update_master_data_only()
