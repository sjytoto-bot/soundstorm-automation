#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SOUNDSTORM 폴더 분석 스크립트 v3
간단한 방식: 전체 스캔 → 필터링
"""

import os
import hashlib
import re
from pathlib import Path
from datetime import datetime
from collections import defaultdict
import warnings
warnings.filterwarnings('ignore')

import pandas as pd
import numpy as np
from tqdm import tqdm
import librosa
from mutagen import File as MutagenFile
from sklearn.metrics.pairwise import cosine_similarity
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from core.path_config import PROJECT_ROOT

# 설정
BASE_DIR = PROJECT_ROOT

# 제외할 패턴 (경로에 포함되면 제외)
EXCLUDED_PATH_PATTERNS = [
    'Audio Files',
    '.venv',
    '__pycache__',
    'node_modules',
    '.git',
    'site-packages',
]

# 제외할 파일 확장자
EXCLUDED_EXTENSIONS = {'.py', '.json', '.txt', '.md', '.csv', '.gsheet', '.html'}

# 분석 대상 파일 확장자
AUDIO_EXTENSIONS = {'.wav', '.mp3', '.aiff', '.flac', '.m4a', '.aac', '.ogg'}
IMAGE_EXTENSIONS = {'.png', '.jpg', '.jpeg', '.gif', '.bmp', '.tiff'}
VIDEO_EXTENSIONS = {'.mp4', '.mov', '.avi', '.mkv'}
DOCUMENT_EXTENSIONS = {'.pdf', '.docx', '.doc'}
PROJECT_EXTENSIONS = {'.psxprj'}


def should_exclude_file(file_path):
    """파일을 제외해야 하는지 확인"""
    # 경로에 제외 패턴이 있는지 확인
    for pattern in EXCLUDED_PATH_PATTERNS:
        if pattern in file_path:
            return True
    
    # 숨김 파일 제외
    if os.path.basename(file_path).startswith('.'):
        return True
    
    # 제외 확장자 체크
    ext = Path(file_path).suffix.lower()
    if ext in EXCLUDED_EXTENSIONS:
        return True
    
    # 0KB 파일 제외
    try:
        if os.path.getsize(file_path) == 0:
            return True
    except:
        return True
    
    return False


def get_file_type(file_path):
    """파일 타입 분류"""
    ext = Path(file_path).suffix.lower()
    if ext in AUDIO_EXTENSIONS:
        return 'audio'
    elif ext in IMAGE_EXTENSIONS:
        return 'image'
    elif ext in VIDEO_EXTENSIONS:
        return 'video'
    elif ext in DOCUMENT_EXTENSIONS:
        return 'document'
    elif ext in PROJECT_EXTENSIONS:
        return 'project'
    else:
        return 'other'


def calculate_md5(file_path, quick_mode=True):
    """파일의 MD5 해시값 계산"""
    hash_md5 = hashlib.md5()
    try:
        file_size = os.path.getsize(file_path)
        
        # 10MB 이상 파일은 빠른 모드로
        if quick_mode and file_size > 10 * 1024 * 1024:
            with open(file_path, "rb") as f:
                hash_md5.update(f.read(1024 * 1024))
                f.seek(file_size // 2)
                hash_md5.update(f.read(1024 * 1024))
                f.seek(max(0, file_size - 1024 * 1024))
                hash_md5.update(f.read(1024 * 1024))
            return hash_md5.hexdigest() + "_quick"
        else:
            with open(file_path, "rb") as f:
                for chunk in iter(lambda: f.read(65536), b""):
                    hash_md5.update(chunk)
            return hash_md5.hexdigest()
    except Exception as e:
        return None


def scan_all_files():
    """1단계: 전체 파일 스캔"""
    print("📂 Phase 1: 전체 파일 스캔...")
    
    all_files = []
    for root, dirs, files in os.walk(BASE_DIR):
        for file in files:
            file_path = os.path.join(root, file)
            all_files.append(file_path)
    
    print(f"✓ 총 {len(all_files)}개 파일 발견\n")
    return all_files


def filter_files(all_files):
    """2단계: 제외 패턴 필터링"""
    print("🔍 제외 패턴 필터링...")
    
    filtered_files = []
    excluded_count = 0
    
    for file_path in all_files:
        if should_exclude_file(file_path):
            excluded_count += 1
        else:
            filtered_files.append(file_path)
    
    print(f"✓ {excluded_count}개 파일 제외")
    print(f"✓ {len(filtered_files)}개 파일 분석 대상\n")
    
    return filtered_files


def collect_file_info(files):
    """파일 정보 수집"""
    print("📋 파일 정보 수집...")
    
    files_data = []
    for file_path in tqdm(files, desc="정보 수집"):
        try:
            stat = os.stat(file_path)
            file_type = get_file_type(file_path)
            file_size = stat.st_size
            
            # 비디오 파일이나 100MB 이상 파일은 해시 계산 스킵
            skip_hash = file_type == 'video' or file_size > 100 * 1024 * 1024
            
            file_info = {
                'filename': os.path.basename(file_path),
                'path': file_path,
                'relative_path': os.path.relpath(file_path, BASE_DIR),
                'size_bytes': file_size,
                'size_mb': round(file_size / (1024 * 1024), 2),
                'type': file_type,
                'extension': Path(file_path).suffix.lower(),
                'created': datetime.fromtimestamp(stat.st_birthtime).strftime('%Y-%m-%d %H:%M:%S'),
                'modified': datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S'),
                'md5': None if skip_hash else calculate_md5(file_path)
            }
            files_data.append(file_info)
        except Exception as e:
            pass
    
    print(f"✓ {len(files_data)}개 파일 정보 수집 완료\n")
    return files_data


def find_duplicates(files_data):
    """중복 파일 탐지"""
    print("🔍 Phase 2: 중복 파일 탐지...")
    
    hash_groups = defaultdict(list)
    for file_info in files_data:
        if file_info['md5']:
            hash_groups[file_info['md5']].append(file_info)
    
    exact_duplicates = []
    for hash_val, files in hash_groups.items():
        if len(files) > 1:
            files_sorted = sorted(files, key=lambda x: x['created'])
            exact_duplicates.append({
                'original': files_sorted[0],
                'duplicates': files_sorted[1:],
                'hash': hash_val,
                'count': len(files),
                'wasted_space_mb': sum(f['size_mb'] for f in files_sorted[1:])
            })
    
    print(f"✓ {len(exact_duplicates)}개 중복 그룹 발견\n")
    return exact_duplicates


def analyze_audio(files_data):
    """Phase 3: 오디오 메타데이터 분석"""
    print("🎵 Phase 3: 오디오 분석...")
    
    audio_files = [f for f in files_data if f['type'] == 'audio']
    print(f"  - 오디오 파일: {len(audio_files)}개")
    
    metadata = []
    for file_info in tqdm(audio_files[:100], desc="메타데이터 추출"):  # 처음 100개만
        try:
            file_path = file_info['path']
            y, sr = librosa.load(file_path, sr=None, duration=30)
            duration = librosa.get_duration(y=y, sr=sr)
            
            tempo, _ = librosa.beat.beat_track(y=y, sr=sr)
            bpm = float(tempo) if isinstance(tempo, (int, float, np.number)) else None
            
            audio = MutagenFile(file_path)
            bitrate = None
            if audio and hasattr(audio.info, 'bitrate'):
                bitrate = audio.info.bitrate // 1000
            
            metadata.append({
                'filename': file_info['filename'],
                'duration_sec': round(duration, 2),
                'sample_rate': sr,
                'bitrate_kbps': bitrate,
                'bpm': round(bpm, 1) if bpm else None
            })
            file_info['audio_metadata'] = metadata[-1]
        except:
            pass
    
    print(f"✓ {len(metadata)}개 오디오 메타데이터 추출 완료\n")
    return metadata


def generate_report(files_data, duplicates, output_path):
    """Phase 4: Excel 리포트 생성"""
    print("📊 Phase 4: Excel 리포트 생성...")
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # 시트1: 전체목록
        df = pd.DataFrame(files_data)
        df = df[['filename', 'relative_path', 'size_mb', 'type', 'extension', 'created', 'modified', 'md5']]
        df.columns = ['파일명', '경로', '크기(MB)', '타입', '확장자', '생성일', '수정일', 'MD5해시']
        df.to_excel(writer, sheet_name='전체목록', index=False)
        
        # 시트2: 중복파일
        if duplicates:
            rows = []
            for dup_group in duplicates:
                original = dup_group['original']
                for dup in dup_group['duplicates']:
                    rows.append({
                        '원본파일': original['filename'],
                        '원본경로': original['relative_path'],
                        '중복파일': dup['filename'],
                        '중복경로': dup['relative_path'],
                        '파일크기(MB)': dup['size_mb']
                    })
            df = pd.DataFrame(rows)
            df.to_excel(writer, sheet_name='중복파일', index=False)
        else:
            pd.DataFrame({'메시지': ['중복 파일이 없습니다']}).to_excel(writer, sheet_name='중복파일', index=False)
    
    # 스타일 적용
    wb = openpyxl.load_workbook(output_path)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        ws.auto_filter.ref = ws.dimensions
    
    wb.save(output_path)
    print(f"✓ 리포트 생성 완료: {output_path}\n")


def main():
    """메인 실행"""
    print("=" * 60)
    print("🎼 SOUNDSTORM 폴더 분석 v3 (간단한 방식)")
    print("=" * 60)
    print()
    
    # 1단계: 전체 스캔
    all_files = scan_all_files()
    
    # 2단계: 필터링
    filtered_files = filter_files(all_files)
    
    # 파일 정보 수집
    files_data = collect_file_info(filtered_files)
    
    # 통계
    type_counts = defaultdict(int)
    total_size = 0
    for f in files_data:
        type_counts[f['type']] += 1
        total_size += f['size_mb']
    
    print("📊 파일 타입별 통계:")
    for file_type, count in sorted(type_counts.items()):
        print(f"  - {file_type}: {count}개")
    print(f"  - 총 용량: {round(total_size, 2)} MB\n")
    
    # 중복 탐지
    duplicates = find_duplicates(files_data)
    
    # 오디오 분석
    audio_metadata = analyze_audio(files_data)
    
    # 리포트 생성
    output_dir = BASE_DIR / "07_AUTOMATION_자동화" / "data" / "reports"
    output_dir.mkdir(parents=True, exist_ok=True)
    output_file = output_dir / "SOUNDSTORM_analysis_report_v3.xlsx"
    generate_report(files_data, duplicates, output_file)
    
    # 요약
    print("=" * 60)
    print("✅ 분석 완료!")
    print("=" * 60)
    print(f"📁 총 파일: {len(files_data)}개")
    print(f"🔄 중복 파일: {len(duplicates)}개 그룹")
    print(f"🎵 오디오 파일: {len([f for f in files_data if f['type'] == 'audio'])}개")
    print(f"📊 리포트: {output_file}")
    print("=" * 60)


if __name__ == "__main__":
    main()
